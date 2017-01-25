from django.shortcuts import render
import re
import psexec
import wmiexec
import csv
import os
import sys
import sys
import argparse
import socket
import whois
import re
import urllib
import urllib2
from urllib2 import urlopen
from links import LINKS
import emailprotectionslib.spf as spf
import emailprotectionslib.dmarc as dmarc
from models import Machine, Software
from openpyxl import Workbook
from django.views.static import serve
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from django.core.exceptions import ObjectDoesNotExist
from tables import PostTableSoftware



def principal_view(request):
    return render(request, '../templates/blank.html', {})

def spam(request):
    return render(request, '../templates/spam.html', {})

#Eliminar los print de Debug
def resultados_spam(request):
    target = request.POST['ip']

    #NSLOOKUP Y WHOIS
    nslookup = "No info available"
    hostname = "No info available"
    aliaslist = "No info available"
    ipaddrlist = "No info available"
    org = "No info available"
    country = "No info available"
    date = "No info available"
    try:
        host = socket.gethostbyaddr(target)
        nslookup = str(host)
        hostname = str(host[0])
        aliaslist = str(host[1])
        ipaddrlist = str(host[2])
        print "nslookup: "+ str(host[0]) + "\n"
        try:
            w = whois.whois(host[0])
            print "[*] WHOIS info:"
            print "Organization: " + str(w.org)
            org = str(w.org)
            print "Country: " + str(w.country)
            country = str(w.country)
            print "Creation date: " + str(w.creation_date)
            date = str(w.creation_date)
        except:
            print "fallo whois"
            pass
    except:
        print "fallo nslookup"
        pass

    #SPF
    print "[*] Checking SPF record:"
    poor_configuration = True
    spf_configuration = "No info available"
    spf_all = False
    no_spf = False
    spf_all_valor = "No info available"
    try:
        spf_host = spf.SpfRecord.from_domain(target)
        if spf_host is not None and spf_host.record is not None:
            print "Domain " + str(target) + " has SPF record:"
            print "\t" + str(spf_host)
            if spf_host.all_string is not None:
                if spf_host.all_string != "~all" and spf_host.all_string != "-all":
                    print "Host SPF record is weak or has poor configuration: " + str(spf_host.all_string)
                    poor_configuration = True
                    spf_configuration = str(spf_host)
                    spf_all = False
                else:
                    print "Host has " + "GOOD" + " configuration of SPF record. Contains All item: " + str(
                        spf_host.all_string) + "\n"
                    poor_configuration = False
                    spf_configuration = str(spf_host)
                    spf_all = True
                    spf_all_valor = str(spf_host.all_string)

            else:
                print "Host SPF not contains All item"
                spf_all = False

        else:
            no_spf = True
            print str(host) + " has no SPF record!"
            print "It's possible to Spoof domain!"

    except:
        no_spf = True
        print "except"

    #BLACKLISTS
    total_finds = 0
    i = 0
    fuentes = [False, False, False, False, False, False, False, False, False, False, False, False, False, False, False]
    for link, text in LINKS:
        print link
        print text
        try:
            req = urllib2.Request(link)
            oreq = urllib2.build_opener().open(req)
            html = oreq.read()
            http_code = oreq.code
            locate = re.findall('\\b'+target+'\\b', html)
            if http_code == 200 and locate:
                print "listed"
                total_finds += 1
                fuentes[i] = True

        except:
            print "except"

        i += 1

    if total_finds >= 1:
        print "Listed in: " + str(total_finds) + " public Blacklists \n"
    else:
        print "Not listed in public Blacklists"
        print "\n"

    return render(request, '../templates/spam_results.html', {'target': target, 'nslookup': nslookup, 'hostname': hostname, 'aliaslist': aliaslist, 'ipaddrlist': ipaddrlist,
                                                              'poor_configuration': poor_configuration,'spf_configuration': spf_configuration, 'spf_all': spf_all, 'no_spf': no_spf, 'spf_all_valor':spf_all_valor,
                                                              'total': total_finds, 'org': org, 'country': country, 'date': date ,'tor': fuentes[0], 'blacklist': fuentes[1],
                                                              'EmergingThreats': fuentes[2],'AlienVault': fuentes[3],'BlocklistDE': fuentes[4],'DragonSSH': fuentes[5],
                                                              'DragonVNC': fuentes[6],'OpenBLock': fuentes[7],'NoThinkMalware': fuentes[8],'NoThinkSSH': fuentes[9],
                                                              'Feodo': fuentes[10],'antispam': fuentes[11],'dshield': fuentes[12],'malc0de': fuentes[13],'MalWareBytes': fuentes[14]})

def software(request):
    hosts = Machine.objects.all()
    soft = Software.objects.all()
    return render(request, '../templates/software.html', {'hosts': hosts, 'soft': soft})

def software_detail(request,ip):
    ipb = ip
    print ipb
    if Machine.objects.filter(ip=ipb).exists():
        host = Machine.objects.get(ip=ipb)
        soft = host.software_installed.all()
        return render(request, '../templates/software_detail.html', {'ip': ipb, 'soft': soft})
    elif Machine.objects.filter(hostname=ipb).exists():
        host = Machine.objects.get(hostname=ipb)
        soft = host.software_installed.all()
        return render(request, '../templates/software_detail.html', {'ip': ipb, 'soft': soft})

def software_host_detail(request, soft):
    softw = Software.objects.get(name=soft)
    hosts = Machine.objects.filter(software_installed=softw)
    return render(request, '../templates/software_host_detail.html',{'hosts': hosts, 'soft': softw})

# Proceso para sacar el software de las maquinas
def sacar_software(request):
    print "[LOG]: Inicio proceso, sacando software..."

    #Provisional, no dejarlas hardcodeadas.
    ADMU = 'ADMINST_AAE'
    ADMP = '16OTSI2016otsi1'

    '''
    obj = psexec.PSEXEC('reg query HKLM\Software\Microsoft\Windows\CurrentVersion\Uninstall /s /t REG_SZ', None, None,
                        None, '445', ADMU, ADMP, 'AD', None, None, False, None)
    executer = PSEXEC(command, options.path, options.file, options.c, int(options.port), username, password, domain,
                      options.hashes, options.aesKey, options.k, options.dc_ip)
e   xecuter = wmiexec.WMIEXEC('ipconfig', 'ADMINST_AAE', '16OTSI2016otsi1', 'AD',None, None,
                           None, False, False, None)

    obj = psexec.PSEXEC('reg query HKLM\Software\Microsoft\Windows\CurrentVersion\Uninstall /s /t REG_SZ', None,
                                None, None, '445', ADMU, ADMP, 'AD', None, None, False, None)
    '''

    #Bucle con:
    for host in Machine.objects.all():
        print "Extrayendo software del host: " + host.hostname

        #Comprovar si la maquina esta activa
        active = check_active(host.hostname)
        if active:
            # if comprobacion que la maquina da ping...:
            '''obj = wmiexec.WMIEXEC('reg query HKLM\Software\Microsoft\Windows\CurrentVersion\Uninstall /s /t REG_SZ', ADMU,
                            ADMP, 'AD', None, None, None, False, False, None)'''
            try:
                print "[LOG] Running wmiexec.."
                #obj.run(host.ip)
                command  = 'python SoftwareMonitoring\wmiexec.py AD/ADMINST_AAE:16OTSI2016otsi1@host reg query HKLM\Software\Microsoft\Windows\CurrentVersion\Uninstall /s /t REG_SZ'
                command2 = 'python SoftwareMonitoring\wmiexec.py AD/ADMINST_AAE:16OTSI2016otsi1@host reg query HKLM\Software\Microsoft\Windows\CurrentVersion\Uninstall /s /t REG_SZ'
                h = host.hostname.rstrip()
                com = re.sub('host', h, command2)
                print '[LOG] Commando utilizado: ' + com
                os.system(com)
                print "[LOG] Acabo WMIEXEC, empezamos parsing..."
                print "[LOG] Abirmos fichero para parsear"
                reg = open("salida_reg.txt", "r")

                num_software = 0
                nombre_actual = ''
                version_actual = ''
                publisher_actual = ''

                #Parser
                for line in reg:
                    if re.search(r'\bHKEY', line):
                        if num_software != 0:
                            # Guardar la info en la BD
                            print nombre_actual + " : " + version_actual
                            #if check_bd(nombre_actual,version_actual):
                            if check_bd(nombre_actual,version_actual):
                                host.software_installed.add(Software.objects.filter(name=nombre_actual, version=version_actual))
                            else:
                                soft = Software.objects.create(name=nombre_actual, version=version_actual, publisher=publisher_actual)
                                host.software_installed.add(soft)

                        num_software += 1

                    else:
                        if re.search(r'\bDisplayName\b', line):
                            valores = line.split()
                            nombre_actual_dirty = " ".join(valores[2:])
                            nombre_actual = re.sub(r'[^\x00-\x7f]',r'', nombre_actual_dirty)

                        elif re.search(r'\bDisplayVersion\b', line):
                            valores = line.split()
                            version_actual_dirty = " ".join(valores[2:])
                            version_actual = re.sub(r'[^\x00-\x7f]',r'',version_actual_dirty)

                        elif re.search(r'\bPublisher\b', line):
                            valores = line.split()
                            publisher_actual_dirty = " ".join(valores[2:])
                            publisher_actual = re.sub(r'[^\x00-\x9f]',r'',publisher_actual_dirty)

                # Guardar la info en la BD
                print nombre_actual + " : " + version_actual
                #if check_bd(nombre_actual, version_actual):
                if check_bd(nombre_actual,version_actual):
                    host.software_installed.add(Software.objects.filter(name=nombre_actual, version=version_actual))
                else:
                    soft = Software.objects.create(name=nombre_actual, version=version_actual, publisher=publisher_actual)
                    host.software_installed.add(soft)

                reg.close()
                #Para no comenter errores posteriormente se borra el contenido del fichero:
                file = open("salida_reg.txt", "w")
                file.write("empty")
                file.close()
                print "[LOG] Fin de la extraccion al host: " + host.hostname

            except:
                file = open("salida_reg.txt", "w")
                file.write("empty")
                file.close()
                print "[LOG] Se ha producido un error en wmiexec o en el parser en la maquina: " + str(host.hostname)

    return render(request, '../templates/software.html', {'success': True})

def do_xls(request):
    #create xls, in future will be a Jaume xls template
    wb = Workbook()
    ws = wb.active
    ws.title = "Software in autopistas"
    ws['A1'] = "HOST"
    ws.column_dimensions['A'].width = 15
    ws['B1'] = "Software"
    ws.column_dimensions['B'].width = 15
    ws['C1'] = "Version"
    ws.column_dimensions['B'].width = 15

    whiteFont = Font(name='Arial',
                     size=10,
                     color='FFFFFFFF')

    blueFill = PatternFill(start_color='0066CC',
                           end_color='0066CC',
                           fill_type='solid')

    border_negro = Border(left=Side(border_style='thin', color='FF000000'),
                          right=Side(border_style='thin', color='FF000000'),
                          top=Side(border_style='thin', color='FF000000'),
                          bottom=Side(border_style='thin', color='FF000000')
                          )

    wrap_text = Alignment(horizontal='general',
                          vertical='bottom',
                          text_rotation=0,
                          wrap_text=True,
                          shrink_to_fit=False,
                          indent=0)

    ## Header:
    for k in range(1, 4):
        ws.cell(row=1, column=k).font = whiteFont
        ws.cell(row=1, column=k).fill = blueFill
        ws.cell(row=1, column=k).alignment = wrap_text
        ws.cell(row=1, column=k).border = border_negro

    # Hacer el bucle con todos los hosts
    i = 2
    for host in Machine.objects.all():
        for software in host.software_installed.all():
            ws.cell(row=i, column=1).value = host.hostname
            ws.cell(row=i, column=2).value = software.name
            ws.cell(row=i, column=3).value = software.version
            i += 1

    wb.save('software_xls.xls')
    print "[DEBUG ALEX]: xls with all software ready!"
    filepath = 'software_xls.xls'
    return serve(request, os.path.basename(filepath), os.path.dirname(filepath))

def do_csv(request):
    csv.register_dialect(
        'standar',
        delimiter=';',
        quotechar='"',
        doublequote=False,
        skipinitialspace=True,
        lineterminator='\n',
        quoting=csv.QUOTE_MINIMAL)

    with open('software_csv.csv', 'w') as file:
        datawriter = csv.writer(file, dialect='standar')
        for host in Machine.objects.all():
            for software in host.software_installed.all():
                soft = str(software.name)
                array = [host.hostname, soft, software.version]
                datawriter.writerow(array)

    filepath = 'software_csv.csv'
    return serve(request, os.path.basename(filepath), os.path.dirname(filepath))

def check_active(hostname):
    if sys.platform == "win32":
        response = os.system("ping " + hostname)
    else:
        response = os.system("ping -c 1 " + hostname)

    if response == 0:
        pingstatus = True
    else:
        pingstatus = False

    return pingstatus

def insertar_maquinas(request):
    if request.method == 'POST' and request.FILES['file']:
        try:
            if request.POST['ip']:
                myfile = request.FILES['file']
                for line in myfile:
                    Machine.objects.create(ip=str(line))
        except:
            pass

        try:
            if request.POST['host']:
                myfile = request.FILES['file']
                for line in myfile:
                    Machine.objects.create(hostname=str(line),ip=str(line))
        except:
            pass

    return render(request, '../templates/software.html', {'success': True})

#Devulve un bool True si existe, false si no existe: es posible con except .exists()
def check_bd(nombre_actual, version_actual):
    return False
    '''if Software.objects.filter(name=nombre_actual, version=version_actual).exists():
        return True
    else:
        return False'''

def datatable(request):
    posts = PostTableSoftware()
    return render(request, "datatable.html", {'posts': posts})